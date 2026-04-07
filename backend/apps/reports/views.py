from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from django.db.models import Sum, Count, F, Q, Avg
from django.db.models.functions import TruncDate, TruncMonth, TruncWeek
from django.utils import timezone
from django.http import HttpResponse
from datetime import timedelta, date
import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from apps.billing.models import Invoice, InvoiceItem
from apps.accounting.models import Expense
from apps.inventory.models import Product


class DashboardView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        today = datetime.date.today()
        month_start = today.replace(day=1)

        today_sales = Invoice.objects.filter(
            created_at__date=today, status=Invoice.STATUS_PAID
        ).aggregate(total=Sum('total_amount'), count=Count('id'))

        month_sales = Invoice.objects.filter(
            created_at__date__gte=month_start, status=Invoice.STATUS_PAID
        ).aggregate(total=Sum('total_amount'), count=Count('id'))

        today_expenses = Expense.objects.filter(date=today).aggregate(total=Sum('amount'))

        low_stock_count = Product.objects.filter(
            stock_quantity__lte=F('low_stock_threshold'), is_active=True
        ).count()

        # Last 7 days sales trend
        week_ago = today - timedelta(days=6)
        daily_trend = Invoice.objects.filter(
            created_at__date__gte=week_ago, status=Invoice.STATUS_PAID
        ).annotate(date=TruncDate('created_at')).values('date').annotate(
            total=Sum('total_amount'), count=Count('id')
        ).order_by('date')

        return Response({
            'today_sales': today_sales,
            'month_sales': month_sales,
            'today_expenses': today_expenses,
            'low_stock_count': low_stock_count,
            'daily_trend': list(daily_trend),
        })


class SalesReportView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        period = request.query_params.get('period', 'daily')
        start = request.query_params.get('start')
        end = request.query_params.get('end')

        qs = Invoice.objects.filter(status=Invoice.STATUS_PAID)
        if start:
            qs = qs.filter(created_at__date__gte=start)
        if end:
            qs = qs.filter(created_at__date__lte=end)

        trunc_fn = {'daily': TruncDate, 'weekly': TruncWeek, 'monthly': TruncMonth}.get(period, TruncDate)

        data = qs.annotate(period=trunc_fn('created_at')).values('period').annotate(
            total_sales=Sum('total_amount'),
            total_discount=Sum('discount_amount'),
            total_gst=Sum('gst_amount'),
            invoice_count=Count('id'),
        ).order_by('period')

        summary = qs.aggregate(
            total=Sum('total_amount'),
            discount=Sum('discount_amount'),
            gst=Sum('gst_amount'),
            count=Count('id'),
        )

        return Response({'data': list(data), 'summary': summary})


class ProductPerformanceView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        start = request.query_params.get('start')
        end = request.query_params.get('end')
        limit = int(request.query_params.get('limit', 20))

        qs = InvoiceItem.objects.filter(invoice__status=Invoice.STATUS_PAID)
        if start:
            qs = qs.filter(invoice__created_at__date__gte=start)
        if end:
            qs = qs.filter(invoice__created_at__date__lte=end)

        top_products = qs.values('product__name', 'product_id').annotate(
            total_qty=Sum('quantity'),
            total_revenue=Sum('total_price'),
            order_count=Count('invoice', distinct=True),
        ).order_by('-total_revenue')[:limit]

        return Response({'top_products': list(top_products)})


class ProfitLossView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        start = request.query_params.get('start')
        end = request.query_params.get('end')

        inv_qs = Invoice.objects.filter(status=Invoice.STATUS_PAID)
        exp_qs = Expense.objects.all()

        if start:
            inv_qs = inv_qs.filter(created_at__date__gte=start)
            exp_qs = exp_qs.filter(date__gte=start)
        if end:
            inv_qs = inv_qs.filter(created_at__date__lte=end)
            exp_qs = exp_qs.filter(date__lte=end)

        sales = inv_qs.aggregate(
            revenue=Sum('total_amount'),
            discount=Sum('discount_amount'),
            gst=Sum('gst_amount'),
        )

        # COGS from purchase prices
        cogs = InvoiceItem.objects.filter(
            invoice__in=inv_qs
        ).aggregate(
            total_cost=Sum(F('quantity') * F('product__purchase_price'))
        )

        expenses = exp_qs.aggregate(total=Sum('amount'))

        revenue = float(sales['revenue'] or 0)
        cost = float(cogs['total_cost'] or 0)
        expense_total = float(expenses['total'] or 0)
        gross_profit = revenue - cost
        net_profit = gross_profit - expense_total

        return Response({
            'revenue': revenue,
            'cogs': cost,
            'gross_profit': gross_profit,
            'expenses': expense_total,
            'net_profit': net_profit,
            'gross_margin': round((gross_profit / revenue * 100) if revenue else 0, 2),
        })


class ExportExcelView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        report_type = request.query_params.get('type', 'sales')
        start = request.query_params.get('start')
        end = request.query_params.get('end')

        wb = openpyxl.Workbook()
        ws = wb.active
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='1a56db', end_color='1a56db', fill_type='solid')

        if report_type == 'sales':
            ws.title = 'Sales Report'
            headers = ['Invoice No', 'Date', 'Customer', 'Subtotal', 'Discount', 'GST', 'Total', 'Payment', 'Status']
            qs = Invoice.objects.select_related('customer')
            if start:
                qs = qs.filter(created_at__date__gte=start)
            if end:
                qs = qs.filter(created_at__date__lte=end)
            rows = qs.values_list(
                'invoice_number', 'created_at', 'customer__name',
                'subtotal', 'discount_amount', 'gst_amount', 'total_amount',
                'payment_method', 'status'
            )
        elif report_type == 'inventory':
            ws.title = 'Inventory Report'
            headers = ['SKU', 'Product', 'Category', 'Stock', 'Purchase Price', 'Selling Price', 'Low Stock']
            rows = Product.objects.select_related('category').values_list(
                'sku', 'name', 'category__name', 'stock_quantity',
                'purchase_price', 'selling_price', 'low_stock_threshold'
            )
        else:
            rows = []
            headers = []

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        for row_idx, row in enumerate(rows, 2):
            for col_idx, value in enumerate(row, 1):
                ws.cell(row=row_idx, column=col_idx, value=str(value) if value else '')

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{report_type}_report.xlsx"'
        wb.save(response)
        return response


class SettingsView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        from django.conf import settings as django_settings
        return Response({
            'store_name':    getattr(django_settings, 'STORE_NAME', ''),
            'store_address': getattr(django_settings, 'STORE_ADDRESS', ''),
            'store_phone':   getattr(django_settings, 'STORE_PHONE', ''),
            'store_gst':     getattr(django_settings, 'STORE_GST', ''),
        })

    def post(self, request):
        """Update .env file with new store settings"""
        import os, re
        env_path = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))), '.env')
        fields = {
            'STORE_NAME':    request.data.get('store_name', ''),
            'STORE_ADDRESS': request.data.get('store_address', ''),
            'STORE_PHONE':   request.data.get('store_phone', ''),
            'STORE_GST':     request.data.get('store_gst', ''),
        }
        try:
            with open(env_path, 'r', encoding='utf-8') as f:
                content = f.read()
            for key, value in fields.items():
                pattern = rf'^{key}=.*$'
                replacement = f'{key}={value}'
                if re.search(pattern, content, re.MULTILINE):
                    content = re.sub(pattern, replacement, content, flags=re.MULTILINE)
                else:
                    content += f'\n{key}={value}'
            with open(env_path, 'w', encoding='utf-8') as f:
                f.write(content)
            return Response({'success': True, 'message': 'Settings saved. Restart server to apply.'})
        except Exception as e:
            return Response({'error': str(e)}, status=400)


class LowStockView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        products = Product.objects.filter(
            stock_quantity__lte=F('low_stock_threshold'), is_active=True
        ).select_related('category').values(
            'id', 'name', 'sku', 'stock_quantity', 'low_stock_threshold', 'category__name'
        ).order_by('stock_quantity')
        return Response({'count': products.count(), 'products': list(products)})


class CustomerHistoryView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, customer_id):
        from apps.accounting.models import Customer, CustomerPayment
        from apps.billing.serializers import InvoiceListSerializer
        try:
            customer = Customer.objects.get(id=customer_id)
        except Customer.DoesNotExist:
            return Response({'error': 'Not found'}, status=404)

        invoices = Invoice.objects.filter(customer=customer).order_by('-created_at')[:20]
        payments = CustomerPayment.objects.filter(customer=customer).order_by('-created_at')[:10]

        total_spent = invoices.aggregate(t=Sum('total_amount'))['t'] or 0
        visit_count = invoices.count()

        return Response({
            'customer': {
                'id': customer.id, 'name': customer.name,
                'phone': customer.phone, 'email': customer.email,
                'outstanding_balance': float(customer.outstanding_balance),
            },
            'total_spent': float(total_spent),
            'visit_count': visit_count,
            'invoices': InvoiceListSerializer(invoices, many=True).data,
            'payments': list(payments.values('amount', 'payment_method', 'created_at', 'notes')),
        })


class ExpiryTrackingView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        from apps.inventory.models import Batch
        import datetime
        today = datetime.date.today()
        days = int(request.query_params.get('days', 30))
        threshold = today + datetime.timedelta(days=days)

        expired = Batch.objects.filter(
            expiry_date__lt=today, quantity__gt=0
        ).select_related('product').values(
            'id', 'product__name', 'product__sku', 'batch_number',
            'expiry_date', 'quantity'
        ).order_by('expiry_date')

        expiring_soon = Batch.objects.filter(
            expiry_date__gte=today, expiry_date__lte=threshold, quantity__gt=0
        ).select_related('product').values(
            'id', 'product__name', 'product__sku', 'batch_number',
            'expiry_date', 'quantity'
        ).order_by('expiry_date')

        return Response({
            'expired': list(expired),
            'expiring_soon': list(expiring_soon),
            'expired_count': len(expired),
            'expiring_soon_count': len(expiring_soon),
        })


class DashboardQuickStatsView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        from apps.accounting.models import Customer
        from apps.ecommerce.models import OnlineOrder
        from apps.billing.models import Invoice
        import datetime

        total_customers = Customer.objects.filter(is_active=True).count()
        total_products  = Product.objects.filter(is_active=True).count()
        total_revenue   = Invoice.objects.filter(status='paid').aggregate(t=Sum('total_amount'))['t'] or 0

        recent_orders = []
        try:
            recent_orders = [{
                'order_number': o.order_number,
                'customer_name': o.customer.name if o.customer else '—',
                'total_amount': float(o.total_amount),
                'status': o.status,
                'created_at': o.created_at,
            } for o in OnlineOrder.objects.select_related('customer').order_by('-created_at')[:5]]
        except Exception:
            pass

        recent_invoices = [{
            'invoice_number': i.invoice_number,
            'customer_name': i.customer.name if i.customer else 'Walk-in',
            'total_amount': float(i.total_amount),
            'payment_method': i.payment_method,
            'created_at': i.created_at,
        } for i in Invoice.objects.select_related('customer').order_by('-created_at')[:5]]

        return Response({
            'total_customers': total_customers,
            'total_products':  total_products,
            'total_revenue':   float(total_revenue),
            'recent_orders':   recent_orders,
            'recent_invoices': recent_invoices,
        })


class WhatsAppReminderView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        from apps.accounting.models import Customer
        customer_id = request.data.get('customer_id')
        try:
            customer = Customer.objects.get(id=customer_id)
        except Customer.DoesNotExist:
            return Response({'error': 'Customer not found'}, status=404)

        if not customer.phone:
            return Response({'error': 'Customer has no phone number'}, status=400)

        balance = float(customer.outstanding_balance)
        if balance <= 0:
            return Response({'error': 'No outstanding balance'}, status=400)

        # Try to send via Twilio if configured
        from django.conf import settings as django_settings
        sid   = getattr(django_settings, 'TWILIO_ACCOUNT_SID', '')
        token = getattr(django_settings, 'TWILIO_AUTH_TOKEN', '')
        from_  = getattr(django_settings, 'TWILIO_WHATSAPP_FROM', '')

        if sid and token and from_:
            try:
                from twilio.rest import Client
                client = Client(sid, token)
                msg = f"Dear {customer.name}, your outstanding balance at Sultan Mart is ₹{balance:.2f}. Please clear at your earliest convenience. Thank you!"
                client.messages.create(body=msg, from_=from_, to=f"whatsapp:{customer.phone}")
                return Response({'success': True, 'message': f'WhatsApp reminder sent to {customer.phone}'})
            except Exception as e:
                return Response({'error': str(e)}, status=500)
        else:
            # Return message template if Twilio not configured
            msg = f"Dear {customer.name}, your outstanding balance at Sultan Mart is ₹{balance:.2f}. Please clear at your earliest convenience."
            return Response({'success': False, 'message': 'Twilio not configured', 'template': msg})


class BackupView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        """Download full DB backup as JSON"""
        import json
        from django.core import serializers as dj_serializers
        from apps.billing.models import Invoice, InvoiceItem
        from apps.accounting.models import Customer, Supplier, PurchaseOrder, Expense
        from apps.inventory.models import Product, Category, Batch

        data = {}
        for name, model in [
            ('customers', Customer), ('suppliers', Supplier),
            ('products', Product), ('categories', Category),
            ('invoices', Invoice), ('invoice_items', InvoiceItem),
            ('expenses', Expense), ('batches', Batch),
        ]:
            data[name] = json.loads(dj_serializers.serialize('json', model.objects.all()))

        response = HttpResponse(json.dumps(data, indent=2, default=str), content_type='application/json')
        response['Content-Disposition'] = 'attachment; filename="sultanmart_backup.json"'
        return response


class RestoreView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        """Restore DB from JSON backup file"""
        import json
        from django.core import serializers as dj_serializers
        from django.db import transaction

        backup_file = request.FILES.get('file')
        if not backup_file:
            return Response({'error': 'No file uploaded'}, status=400)

        try:
            data = json.loads(backup_file.read().decode('utf-8'))
        except Exception:
            return Response({'error': 'Invalid JSON file'}, status=400)

        # Restore order matters — categories before products, customers before invoices
        restore_order = ['categories', 'customers', 'suppliers', 'products',
                         'batches', 'expenses', 'invoices', 'invoice_items']

        restored = {}
        errors = []
        with transaction.atomic():
            for key in restore_order:
                if key not in data:
                    continue
                try:
                    objects = list(dj_serializers.deserialize('json', json.dumps(data[key])))
                    for obj in objects:
                        obj.save()
                    restored[key] = len(objects)
                except Exception as e:
                    errors.append(f'{key}: {str(e)}')

        if errors:
            return Response({'success': False, 'errors': errors}, status=400)

        return Response({'success': True, 'restored': restored})


class VerifyPinView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        from django.conf import settings as django_settings
        pin = request.data.get('pin', '')
        correct = getattr(django_settings, 'ADMIN_PIN', '1234')
        if pin == correct:
            return Response({'valid': True})
        return Response({'valid': False}, status=400)

    def put(self, request):
        """Update PIN in .env file"""
        import os, re
        new_pin = request.data.get('pin', '')
        if not re.match(r'^\d{4}$', new_pin):
            return Response({'error': 'PIN must be 4 digits'}, status=400)
        env_path = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))), '.env')
        try:
            with open(env_path, 'r') as f:
                content = f.read()
            if re.search(r'^ADMIN_PIN=', content, re.MULTILINE):
                content = re.sub(r'^ADMIN_PIN=.*', f'ADMIN_PIN={new_pin}', content, flags=re.MULTILINE)
            else:
                content += f'\nADMIN_PIN={new_pin}'
            with open(env_path, 'w') as f:
                f.write(content)
            return Response({'success': True})
        except Exception as e:
            return Response({'error': str(e)}, status=400)
